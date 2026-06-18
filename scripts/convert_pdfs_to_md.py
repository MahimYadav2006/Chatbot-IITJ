#!/usr/bin/env python3
"""
PDF & Excel → Markdown Converter for IIT Jammu Chatbot
=======================================================
Handles:
  - Digitally-typed PDFs  → pdfplumber text + table extraction
  - Mobile-scanned PDFs   → OCR via pytesseract (pdf2image + tesseract)
  - Excel (.xlsx) files   → openpyxl → markdown tables
  
Architecture:
  1. Classify each file (digital vs scanned vs excel)
  2. Extract text using the best strategy
  3. Detect and render tables as markdown
  4. Post-process: clean up, add metadata headers
  5. Write structured .md files preserving directory hierarchy
"""

import os
import re
import sys
import json
import hashlib
import logging
import argparse
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple, Optional

import pdfplumber
import pytesseract
from pdf2image import convert_from_path
from PIL import Image, ImageFilter, ImageEnhance
import openpyxl

# ── Logging ──────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s │ %(levelname)-7s │ %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("pdf2md")

# ── Constants ────────────────────────────────────────────────────────────
SCANNED_THRESHOLD = 50        # chars per page to consider "scanned"
OCR_DPI = 300                 # DPI for rasterising scanned pages
TABLE_MIN_ROWS = 2            # minimum rows to consider a valid table
MAX_IMAGE_PIXELS = 20_000_000 # safety cap for large page images


# ═══════════════════════════════════════════════════════════════════════
#  CLASSIFICATION
# ═══════════════════════════════════════════════════════════════════════

def classify_pdf(path: str) -> str:
    """Return 'digital', 'scanned', or 'mixed'."""
    try:
        pdf = pdfplumber.open(path)
    except Exception as e:
        log.warning(f"Cannot open {path}: {e}")
        return "scanned"  # fallback to OCR

    digital_pages = 0
    total = len(pdf.pages)
    sample = min(total, 5)  # check first 5 pages

    for page in pdf.pages[:sample]:
        text = (page.extract_text() or "").strip()
        if len(text) > SCANNED_THRESHOLD:
            digital_pages += 1

    pdf.close()
    ratio = digital_pages / sample if sample else 0
    if ratio > 0.8:
        return "digital"
    elif ratio < 0.2:
        return "scanned"
    return "mixed"


# ═══════════════════════════════════════════════════════════════════════
#  TEXT EXTRACTION — DIGITAL PDFs
# ═══════════════════════════════════════════════════════════════════════

def extract_table_as_md(table: list) -> str:
    """Convert a pdfplumber table (list of lists) to markdown table."""
    if not table or len(table) < TABLE_MIN_ROWS:
        return ""
    # Clean cells
    cleaned = []
    for row in table:
        cleaned.append([
            (cell or "").replace("\n", " ").strip() for cell in row
        ])
    if not cleaned:
        return ""

    # Build markdown table
    header = cleaned[0]
    col_count = len(header)
    lines = ["| " + " | ".join(header) + " |"]
    lines.append("| " + " | ".join(["---"] * col_count) + " |")
    for row in cleaned[1:]:
        # Pad row if needed
        while len(row) < col_count:
            row.append("")
        lines.append("| " + " | ".join(row[:col_count]) + " |")
    return "\n".join(lines)


def extract_digital_pdf(path: str) -> str:
    """Extract text + tables from a digitally-typed PDF."""
    pdf = pdfplumber.open(path)
    sections = []

    for i, page in enumerate(pdf.pages):
        page_parts = []

        # Extract tables first
        tables = page.extract_tables() or []
        table_bboxes = []
        table_texts = []

        for tbl in tables:
            md_table = extract_table_as_md(tbl)
            if md_table:
                table_texts.append(md_table)

        # Extract full page text
        text = (page.extract_text() or "").strip()

        if text:
            page_parts.append(text)

        # Append tables after text
        for t in table_texts:
            page_parts.append("\n" + t + "\n")

        if page_parts:
            content = "\n".join(page_parts)
            sections.append(f"<!-- Page {i+1} -->\n{content}")

    pdf.close()
    return "\n\n---\n\n".join(sections)


# ═══════════════════════════════════════════════════════════════════════
#  TEXT EXTRACTION — SCANNED PDFs (OCR)
# ═══════════════════════════════════════════════════════════════════════

def preprocess_image(img: Image.Image) -> Image.Image:
    """Enhance scanned image for better OCR accuracy."""
    # Convert to grayscale
    img = img.convert("L")
    # Increase contrast
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(1.5)
    # Sharpen
    img = img.filter(ImageFilter.SHARPEN)
    # Binarize with adaptive threshold
    img = img.point(lambda x: 0 if x < 140 else 255, "1")
    return img


def ocr_page_image(img: Image.Image, page_num: int) -> str:
    """Run OCR on a single page image."""
    processed = preprocess_image(img)
    try:
        # Use PSM 6 (assume uniform block of text) for document pages
        custom_config = r"--oem 3 --psm 6"
        text = pytesseract.image_to_string(processed, config=custom_config)
        return text.strip()
    except Exception as e:
        log.error(f"  OCR failed on page {page_num}: {e}")
        return ""


def extract_scanned_pdf(path: str) -> str:
    """Extract text from scanned PDF via OCR."""
    try:
        images = convert_from_path(path, dpi=OCR_DPI)
    except Exception as e:
        log.error(f"  Failed to rasterize {path}: {e}")
        return ""

    sections = []
    for i, img in enumerate(images):
        log.info(f"  OCR page {i+1}/{len(images)}...")
        text = ocr_page_image(img, i + 1)
        if text:
            sections.append(f"<!-- Page {i+1} -->\n{text}")

    return "\n\n---\n\n".join(sections)


# ═══════════════════════════════════════════════════════════════════════
#  TEXT EXTRACTION — MIXED PDFs
# ═══════════════════════════════════════════════════════════════════════

def extract_mixed_pdf(path: str) -> str:
    """Handle PDFs with both digital and scanned pages."""
    pdf = pdfplumber.open(path)
    sections = []

    try:
        images = convert_from_path(path, dpi=OCR_DPI)
    except Exception:
        images = [None] * len(pdf.pages)

    for i, page in enumerate(pdf.pages):
        text = (page.extract_text() or "").strip()

        # If page has enough digital text, use it
        if len(text) > SCANNED_THRESHOLD:
            tables = page.extract_tables() or []
            parts = [text]
            for tbl in tables:
                md_table = extract_table_as_md(tbl)
                if md_table:
                    parts.append("\n" + md_table + "\n")
            sections.append(f"<!-- Page {i+1} -->\n" + "\n".join(parts))
        else:
            # Fall back to OCR for this page
            if i < len(images) and images[i]:
                ocr_text = ocr_page_image(images[i], i + 1)
                if ocr_text:
                    sections.append(f"<!-- Page {i+1} [OCR] -->\n{ocr_text}")

    pdf.close()
    return "\n\n---\n\n".join(sections)


# ═══════════════════════════════════════════════════════════════════════
#  EXCEL EXTRACTION
# ═══════════════════════════════════════════════════════════════════════

def extract_excel(path: str) -> str:
    """Convert Excel file to markdown tables."""
    wb = openpyxl.load_workbook(path, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Filter out completely empty rows
        rows = [r for r in rows if any(c is not None for c in r)]
        if len(rows) < 1:
            continue

        # Determine max columns
        max_cols = max(len(r) for r in rows)

        section = f"### Sheet: {sheet_name}\n\n"

        # Build table
        header = rows[0]
        header_cells = [(str(c) if c is not None else "") for c in header]
        while len(header_cells) < max_cols:
            header_cells.append("")

        lines = ["| " + " | ".join(header_cells) + " |"]
        lines.append("| " + " | ".join(["---"] * max_cols) + " |")

        for row in rows[1:]:
            cells = [(str(c) if c is not None else "") for c in row]
            while len(cells) < max_cols:
                cells.append("")
            # Escape pipes in cell content
            cells = [c.replace("|", "\\|") for c in cells[:max_cols]]
            lines.append("| " + " | ".join(cells) + " |")

        section += "\n".join(lines)
        sections.append(section)

    wb.close()
    return "\n\n---\n\n".join(sections)


# ═══════════════════════════════════════════════════════════════════════
#  POST-PROCESSING
# ═══════════════════════════════════════════════════════════════════════

def clean_text(text: str) -> str:
    """Clean extracted text for chatbot consumption."""
    # Remove excessive whitespace but preserve paragraph breaks
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{4,}", "\n\n\n", text)
    # Fix common OCR artifacts
    text = text.replace("ﬁ", "fi").replace("ﬂ", "fl")
    text = text.replace("'", "'").replace("'", "'")
    text = text.replace(""", '"').replace(""", '"')
    text = text.replace("–", "-").replace("—", "-")
    # Remove null bytes
    text = text.replace("\x00", "")
    return text.strip()


def generate_metadata_header(
    source_path: str,
    doc_type: str,
    category: str,
    subcategory: str,
) -> str:
    """Generate YAML front-matter for the markdown file."""
    filename = os.path.basename(source_path)
    return f"""---
source_file: "{filename}"
source_path: "{source_path}"
document_type: "{doc_type}"
category: "{category}"
subcategory: "{subcategory}"
converted_at: "{datetime.now().isoformat()}"
institution: "Indian Institute of Technology Jammu"
---

"""


def derive_title(filename: str) -> str:
    """Derive a clean title from the filename."""
    name = os.path.splitext(filename)[0]
    # Clean up common patterns
    name = re.sub(r"_+", " ", name)
    name = re.sub(r"\s+", " ", name)
    return name.strip()


# ═══════════════════════════════════════════════════════════════════════
#  MAIN PIPELINE
# ═══════════════════════════════════════════════════════════════════════

def determine_category(path: str) -> Tuple[str, str]:
    """Determine category and subcategory from file path."""
    parts = Path(path).parts
    category = "academics"
    subcategory = "general"

    # Normalize parts to check for academic notifications (handles any directory separators and capitalization/spacing/typos)
    normalized_parts = [p.lower().replace("_", " ").replace("-", " ").strip() for p in parts]

    if any("academic notifications" in p or "academic notiications" in p for p in normalized_parts):
        category = "academic_notifications"
        subcategory = "general"
    elif "Rules and Regulations" in parts:
        category = "rules_and_regulations"
        if "UG" in parts:
            subcategory = "undergraduate"
        elif "PG" in parts:
            if "PhD" in parts:
                subcategory = "phd"
            elif "MTech" in parts:
                subcategory = "mtech"
            elif "MSc" in parts:
                subcategory = "msc"
            else:
                subcategory = "postgraduate"
    elif "academics-specialisation-and-courses" in parts:
        category = "specialisation_and_courses"
        if "UG" in parts:
            if "Specialization" in parts:
                subcategory = "ug_specialization"
            elif "Course Offering Framework" in parts:
                subcategory = "ug_course_framework"
            else:
                subcategory = "undergraduate"
        elif "PG" in parts:
            if "MTech" in parts:
                subcategory = "mtech_programs"
            elif "MSc" in parts:
                subcategory = "msc_programs"
            elif "PhD" in parts:
                subcategory = "phd_programs"
            else:
                subcategory = "postgraduate"

    return category, subcategory


def make_output_path(input_path: str, base_input: str, output_dir: str) -> str:
    """Create output .md path preserving directory structure."""
    rel = os.path.relpath(input_path, base_input)
    name = os.path.splitext(rel)[0] + ".md"
    # Sanitize path components
    name = name.replace(" ", "_")
    return os.path.join(output_dir, name)


def process_file(
    input_path: str,
    base_input: str,
    output_dir: str,
    force: bool = False,
) -> Optional[Dict]:
    """Process a single file and return a report dict."""
    rel_path = os.path.relpath(input_path, base_input)
    output_path = make_output_path(input_path, base_input, output_dir)
    filename = os.path.basename(input_path)
    ext = os.path.splitext(filename)[1].lower()

    # Skip if already processed (unless force)
    if os.path.exists(output_path) and not force:
        log.info(f"⏭  SKIP (exists): {rel_path}")
        return {"file": rel_path, "status": "skipped", "reason": "already exists"}

    log.info(f"📄 Processing: {rel_path}")

    category, subcategory = determine_category(input_path)
    title = derive_title(filename)
    report = {"file": rel_path, "category": category, "subcategory": subcategory}

    try:
        if ext in (".xlsx", ".xls"):
            doc_type = "excel"
            content = extract_excel(input_path)
            report["method"] = "openpyxl"
        elif ext == ".pdf":
            classification = classify_pdf(input_path)
            report["classification"] = classification
            doc_type = "pdf"

            if classification == "digital":
                content = extract_digital_pdf(input_path)
                report["method"] = "pdfplumber"
            elif classification == "scanned":
                log.info(f"  🔍 Scanned PDF detected → OCR")
                content = extract_scanned_pdf(input_path)
                report["method"] = "tesseract_ocr"
            else:
                log.info(f"  🔀 Mixed PDF → hybrid extraction")
                content = extract_mixed_pdf(input_path)
                report["method"] = "hybrid"
        else:
            log.warning(f"  ⚠ Unsupported format: {ext}")
            return {"file": rel_path, "status": "skipped", "reason": f"unsupported: {ext}"}

        # Post-process
        content = clean_text(content)

        if not content or len(content) < 20:
            log.warning(f"  ⚠ Extracted content too short ({len(content)} chars)")
            report["status"] = "warning"
            report["reason"] = "very short content"

        # Build final markdown
        header = generate_metadata_header(rel_path, doc_type, category, subcategory)
        md = f"{header}# {title}\n\n{content}\n"

        # Write output
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(md)

        report["status"] = "success"
        report["output"] = output_path
        report["chars"] = len(content)
        log.info(f"  ✅ → {output_path} ({len(content)} chars)")

    except Exception as e:
        log.error(f"  ❌ FAILED: {e}")
        report["status"] = "error"
        report["error"] = str(e)

    return report


def run_pipeline(input_dir: str, output_dir: str, force: bool = False):
    """Run the full conversion pipeline."""
    log.info("=" * 70)
    log.info("  IIT Jammu Academic Document → Markdown Converter")
    log.info("=" * 70)
    log.info(f"  Input:  {input_dir}")
    log.info(f"  Output: {output_dir}")
    log.info(f"  Force:  {force}")
    log.info("")

    # Collect all files
    files = []
    for root, dirs, filenames in os.walk(input_dir):
        for fn in sorted(filenames):
            ext = os.path.splitext(fn)[1].lower()
            if ext in (".pdf", ".xlsx", ".xls"):
                files.append(os.path.join(root, fn))

    log.info(f"Found {len(files)} files to process")
    log.info("")

    # Process
    reports = []
    for i, fpath in enumerate(files, 1):
        log.info(f"[{i}/{len(files)}] {'─'*50}")
        report = process_file(fpath, input_dir, output_dir, force)
        if report:
            reports.append(report)

    # Summary
    log.info("")
    log.info("=" * 70)
    log.info("  SUMMARY")
    log.info("=" * 70)

    success = [r for r in reports if r.get("status") == "success"]
    errors  = [r for r in reports if r.get("status") == "error"]
    skipped = [r for r in reports if r.get("status") == "skipped"]
    warnings= [r for r in reports if r.get("status") == "warning"]

    log.info(f"  ✅ Success:  {len(success)}")
    log.info(f"  ⚠  Warnings: {len(warnings)}")
    log.info(f"  ❌ Errors:   {len(errors)}")
    log.info(f"  ⏭  Skipped:  {len(skipped)}")

    if errors:
        log.info("")
        log.info("  Failed files:")
        for r in errors:
            log.info(f"    - {r['file']}: {r.get('error', 'unknown')}")

    # Save report
    report_path = os.path.join(output_dir, "_conversion_report.json")
    os.makedirs(output_dir, exist_ok=True)
    with open(report_path, "w") as f:
        json.dump(reports, f, indent=2)
    log.info(f"\n  Report saved: {report_path}")

    return reports


# ═══════════════════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert academic PDFs & Excel files to Markdown"
    )
    parser.add_argument(
        "--input", "-i",
        default="scraped_data/sections/academics/Pdf",
        help="Input directory containing PDFs/Excel files",
    )
    parser.add_argument(
        "--output", "-o",
        default="scraped_data/sections/academics/parsed_documents",
        help="Output directory for .md files",
    )
    parser.add_argument(
        "--force", "-f",
        action="store_true",
        help="Re-process files even if output already exists",
    )
    args = parser.parse_args()

    run_pipeline(args.input, args.output, args.force)
